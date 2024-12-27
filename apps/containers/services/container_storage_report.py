from typing import Any, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from apps.containers.models import ContainerStorage


class TransportType:
    AUTO = "AUTO"
    WAGON = "WAGON"


class ContainerStorageReportService:
    def __init__(self):
        self.headers = {
            "in_terminal": [
                "№",
                "Склад",
                "Номер контейнера",
                "Тип",
                "Тек сост конт-ра",
                "Собственник",
                "Дата прибытия",
                "Номер авто/вагона",
                "Вид транс-та прибытия",
                "Cрок хранения (дней)",
                "Примечания",
            ],
            "dispatched": [
                "№",
                "Склад",
                "Номер контейнера",
                "Тип",
                "Собственник",
                "Дата прибытия",
                "Дата убытия",
                "Номер авто/вагона",
                "Вид транс-та отправки",
                "Срок хранения (дней)",
                "Примечания",
            ],
        }

    def get_report(
        self,
        company_id: int,
        dispatched: str,
        transport_type: Optional[str] = None,
        month: Optional[int] = None,
    ) -> Workbook:
        """Generate container report"""
        # Debug incoming value

        filters = {"company_id": company_id}

        # Check dispatched type and convert if needed

        # Set filter based on dispatched status
        if dispatched == "true":
            filters["exit_time__isnull"] = False
        elif dispatched == "false":
            filters["exit_time__isnull"] = True
        else:
            filters = {"company_id": company_id}
        print(f"Final filters: {filters}")

        if transport_type:
            if dispatched == "true":
                filters["exit_transport_type"] = transport_type
            else:
                filters["transport_type"] = transport_type

        if month:
            if dispatched == "true":
                date_field = "exit_time__month"
            else:
                date_field = "entry_time__month"
            filters[date_field] = month

        # Get data
        print(f"Filters: {filters}")
        containers = ContainerStorage.objects.filter(**filters)

        # Create report
        wb = Workbook()
        ws = wb.active

        # Set headers based on report type
        headers = self.headers["dispatched" if dispatched else "in_terminal"]

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Write data
        for row_idx, container in enumerate(containers, 2):
            data = self._get_row_data(container, dispatched)
            for col_idx, value in enumerate(data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            for cell in column:
                try:
                    max_length = max(max_length, len(str(cell.value or "")))
                except TypeError:
                    continue
            ws.column_dimensions[column[0].column_letter].width = max_length + 2

        return wb

    def _get_row_data(self, container, dispatched: bool) -> List[Any]:
        """Get row data based on report type"""
        storage_days = (
            (container.exit_time - container.entry_time).days
            if container.exit_time and container.entry_time
            else 0
        )

        if dispatched:
            return [
                container.id,
                "MTT",
                container.container.name if container.container else "",
                container.container.size if container.container else "",
                container.container_owner,
                container.entry_time.strftime("%d.%m.%Y %H:%M")
                if container.entry_time
                else "",
                container.exit_time.strftime("%d.%m.%Y %H:%M")
                if container.exit_time
                else "",
                container.transport_number,
                container.exit_transport_type,
                storage_days,
                container.notes,
            ]
        else:
            return [
                container.id,
                "MTT",
                container.container.name if container.container else "",
                container.container.size if container.container else "",
                container.container_state,
                container.container_owner,
                container.entry_time.strftime("%d.%m.%Y %H:%M")
                if container.entry_time
                else "",
                container.transport_number,
                container.transport_type,
                storage_days,
                container.notes,
            ]
